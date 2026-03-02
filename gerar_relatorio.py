from funcoes import carregar_cadastros
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def gerar_relatorio():
    cadastros = carregar_cadastros()

    print(f'dados carregados: {cadastros}')

    if not cadastros:
        print("|Nenhum cadastro encontrado para gerar o relatório!|\n")
        return
    
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Cadastros"
    planilha.append(["Nome", "Idade","Email"])

    for pessoa in cadastros:
        planilha.append([pessoa.get("Nome"), pessoa.get("Idade"), pessoa.get("Email")])
    
    for cell in planilha[1]:
        cell.font = Font(bold=True)
    
    for coluna in planilha.columns:
        max_length = 0
        letra_coluna = get_column_letter(coluna[0].column)

        for cell in coluna:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        planilha.column_dimensions[letra_coluna].width = max_length + 4

    ultima_linha = planilha.max_row
    ultima_coluna = planilha.max_column
    referencia = f"A1:{get_column_letter(ultima_coluna)}{ultima_linha}"

    tabela = Table(displayName="TabelaCadastros", ref=referencia)

    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    tabela.tableStyleInfo = estilo
    planilha.add_table(tabela)

    planilha.freeze_panes = "A2"

    workbook.save("relatorio_cadastros.xlsx")
    print("\n|Relatório gerado com sucesso!|\n")

if __name__ == "__main__":
    gerar_relatorio()